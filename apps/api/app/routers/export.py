# -*- coding: utf-8 -*-
"""
Export-router: PDF (via print-sida) och Excel-nedladdning.
"""
import calendar
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.auth_utils import require_admin_or_schemaansvarig
from app.db_models import UserRow

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.db_models import EmployeeRow, SchedulePeriodRow
from app.engine.schemas import (
    Employee, ContractType, Group, Absence, AbsenceType,
    ScheduleDay, Shift, ShiftSegment, ShiftType,
)

router = APIRouter(prefix="/api/export", tags=["export"])

WEEKLY_HOURS = {
    ContractType.DAGTID:       40.0,
    ContractType.VARIERANDE:   37.0,
    ContractType.KVAL:         30.0,
    ContractType.HELG_FRE_SON: 26.0,
    ContractType.HELG_LOR_MAN: 26.0,
    ContractType.NATT:         34.33,
}

SHIFT_ABBREV = {
    ShiftType.DAG_TIDIG: "06:45",
    ShiftType.DAG:       "DAG",
    ShiftType.KVAL_KORT: "K",
    ShiftType.KVAL_LANG: "KL",
    ShiftType.DELAD_TUR: "DT",
    ShiftType.NATT:      "N",
    ShiftType.OBOKAD:    "OB",
    ShiftType.APT:       "APT",
}

ABSENCE_ABBREV = {
    AbsenceType.SEM:  "SEM",
    AbsenceType.FL:   "FL",
    AbsenceType.TJL:  "TJL",
    AbsenceType.SJUK: "SJK",
}

# Hex-färger per passtyp (utan #)
SHIFT_COLOR = {
    ShiftType.DAG_TIDIG: "1e40af",
    ShiftType.DAG:       "3b82f6",
    ShiftType.KVAL_KORT: "7c3aed",
    ShiftType.KVAL_LANG: "6d28d9",
    ShiftType.DELAD_TUR: "4f46e5",
    ShiftType.NATT:      "1e293b",
    ShiftType.OBOKAD:    "9ca3af",
    ShiftType.APT:       "0d9488",
}
ABSENCE_COLOR = {
    AbsenceType.SEM:  "fb923c",
    AbsenceType.FL:   "fdba74",
    AbsenceType.TJL:  "fbbf24",
    AbsenceType.SJUK: "ef4444",
}
WEEKEND_BG  = "dbeafe"
HEADER_BG   = "f1f5f9"
STAFFING_BG = "f8fafc"


def _thin_border():
    s = Side(style="thin", color="e5e7eb")
    return Border(left=s, right=s, top=s, bottom=s)


def _shift_hours(sd_dict: dict) -> float:
    shift = sd_dict.get("shift")
    if not shift or shift.get("is_unbooked"):
        return 0.0
    total = 0.0
    for seg in shift.get("segments", []):
        try:
            s = datetime.fromisoformat(seg["start_time"])
            e = datetime.fromisoformat(seg["end_time"])
            total += (e - s).total_seconds() / 3600.0
        except Exception:
            pass
    return total


@router.get("/{group}/{year}/{month}/excel")
async def export_excel(
    group: str, year: int, month: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserRow = Depends(require_admin_or_schemaansvarig),
):
    _, last_day = calendar.monthrange(year, month)
    days = list(range(1, last_day + 1))

    # Hämta anställda
    own_emp_rows = (await db.execute(
        select(EmployeeRow).where(EmployeeRow.group_name == group)
    )).scalars().all()
    emp_rows = list(own_emp_rows)

    # Hämta schema
    period_row = (await db.execute(
        select(SchedulePeriodRow).where(
            SchedulePeriodRow.group_name == group,
            SchedulePeriodRow.year == year,
            SchedulePeriodRow.month == month,
        )
    )).scalar_one_or_none()

    own_schedule: list[dict] = period_row.schedule if period_row and period_row.schedule else []

    # Hämta inlånad personal från andra gruppers scheman
    other_stmt = select(SchedulePeriodRow).where(
        SchedulePeriodRow.group_name != group,
        SchedulePeriodRow.year == year,
        SchedulePeriodRow.month == month,
    )
    other_result = await db.execute(other_stmt)
    other_rows = other_result.scalars().all()

    borrowed_days = []
    borrowed_emp_ids = set()
    for r in other_rows:
        if not r.schedule:
            continue
        for sd_dict in r.schedule:
            if sd_dict.get("assigned_group") == group:
                borrowed_days.append(sd_dict)
                borrowed_emp_ids.add(sd_dict["employee_id"])

    # Hämta anställningsrader för de inlånade
    if borrowed_emp_ids:
        borrowed_emp_stmt = select(EmployeeRow).where(EmployeeRow.id.in_(list(borrowed_emp_ids)))
        borrowed_emp_rows = (await db.execute(borrowed_emp_stmt)).scalars().all()
        emp_rows.extend(borrowed_emp_rows)

    schedule = own_schedule + borrowed_days

    # Index: employee_id → date → schedule_day_dict
    sched_idx: dict[str, dict[str, dict]] = {}
    for sd in schedule:
        eid = sd["employee_id"]
        d = sd["date"]
        if eid not in sched_idx:
            sched_idx[eid] = {}
        sched_idx[eid][d] = sd

    # ── Skapa Excel-fil ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    month_name = date(year, month, 1).strftime("%B %Y")
    ws.title = f"{group} {month_name}"

    border = _thin_border()

    # ── Rubrikrad ────────────────────────────────────────────────────────────
    ws.cell(1, 1, f"Schema — {group}").font = Font(bold=True, size=13)
    ws.cell(2, 1, f"{month_name.capitalize()}  ·  Genererat {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="6b7280")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 14

    # ── Kolumner: Namn | dag 1..N | Tim | Mål | Saldo ────────────────────────
    DAY_NAMES = ["Mån", "Tis", "Ons", "Tor", "Fre", "Lör", "Sön"]
    HEADER_ROW = 4

    # Namn-kolumn
    ws.column_dimensions["A"].width = 20
    c = ws.cell(HEADER_ROW, 1, "Namn")
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.border = border
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Dagkolumner
    for i, d in enumerate(days):
        col = i + 2
        d_date = date(year, month, d)
        wd = (d_date.weekday()) % 7  # 0=Mån
        weekend = wd >= 5
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 5.5

        # Dag-nummer
        c1 = ws.cell(HEADER_ROW, col, d)
        c1.font = Font(bold=True, size=8, color="1d4ed8" if weekend else "374151")
        c1.fill = PatternFill("solid", fgColor=WEEKEND_BG if weekend else HEADER_BG)
        c1.border = border
        c1.alignment = Alignment(horizontal="center", vertical="center")

        # Veckodagsnamn (rad under)
        c2 = ws.cell(HEADER_ROW + 1, col, DAY_NAMES[wd])
        c2.font = Font(size=7, color="6b7280")
        c2.fill = PatternFill("solid", fgColor=WEEKEND_BG if weekend else HEADER_BG)
        c2.border = border
        c2.alignment = Alignment(horizontal="center", vertical="center")

    # Tim / Mål / Saldo-kolumner
    extra_cols = ["Tim", "Mål", "Saldo"]
    for j, lbl in enumerate(extra_cols):
        col = len(days) + 2 + j
        c = ws.cell(HEADER_ROW, col, lbl)
        c.font = Font(bold=True, size=8)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.border = border
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(HEADER_ROW + 1, col).fill = PatternFill("solid", fgColor=HEADER_BG)
        ws.cell(HEADER_ROW + 1, col).border = border
        ws.column_dimensions[get_column_letter(col)].width = 7

    ws.row_dimensions[HEADER_ROW].height = 16
    ws.row_dimensions[HEADER_ROW + 1].height = 12

    # ── Personalrader ────────────────────────────────────────────────────────
    DATA_START = HEADER_ROW + 2
    staffing_dag   = {d: 0 for d in days}
    staffing_kval  = {d: 0 for d in days}

    for row_idx, emp_row in enumerate(emp_rows):
        row = DATA_START + row_idx
        ws.row_dimensions[row].height = 15
        total_h = 0.0

        # Namn
        c = ws.cell(row, 1, emp_row.name)
        c.font = Font(size=9)
        c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center")

        for i, d in enumerate(days):
            col = i + 2
            d_date = date(year, month, d)
            wd = d_date.weekday()
            weekend = wd >= 5
            datestr = d_date.isoformat()

            sd = sched_idx.get(emp_row.id, {}).get(datestr)
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=8, bold=True, color="FFFFFF")

            if sd:
                h = _shift_hours(sd)
                total_h += h
                shift = sd.get("shift")
                absence = sd.get("absence")
                if absence:
                    atype = absence.get("absence_type", "")
                    try:
                        at = AbsenceType(atype)
                        label = ABSENCE_ABBREV.get(at, atype.upper())
                        color = ABSENCE_COLOR.get(at, "9ca3af")
                    except ValueError:
                        label, color = atype.upper(), "9ca3af"
                    cell.value = label
                    cell.fill = PatternFill("solid", fgColor=color)
                elif shift:
                    try:
                        st = ShiftType(shift["shift_type"])
                        is_ob = shift.get("is_unbooked", False)
                        label = "OB" if is_ob else SHIFT_ABBREV.get(st, st.value.upper())
                        color = "9ca3af" if is_ob else SHIFT_COLOR.get(st, "6b7280")
                    except ValueError:
                        label, color = "?", "9ca3af"
                    cell.value = label
                    cell.fill = PatternFill("solid", fgColor=color)

                    # Bemanningsräknare
                    if shift and not shift.get("is_unbooked"):
                        stype_str = shift.get("shift_type", "")
                        if stype_str in ("dag", "dag_tidig"):
                            staffing_dag[d] += 1
                        elif stype_str in ("kval_kort", "kval_lang"):
                            staffing_kval[d] += 1
                else:
                    cell.fill = PatternFill("solid", fgColor=WEEKEND_BG if weekend else "FFFFFF")
                    cell.font = Font(size=8)
            else:
                cell.fill = PatternFill("solid", fgColor=WEEKEND_BG if weekend else "FFFFFF")
                cell.font = Font(size=8)

        # Tim / Mål / Saldo
        target_h = WEEKLY_HOURS.get(ContractType(emp_row.contract_type), 37.0) * last_day / 7
        saldo = total_h - target_h

        c_tim = ws.cell(row, len(days) + 2, round(total_h, 1))
        c_tim.font = Font(size=9, bold=True)
        c_tim.border = border
        c_tim.alignment = Alignment(horizontal="right", vertical="center")
        c_tim.number_format = "0.0"

        c_mal = ws.cell(row, len(days) + 3, round(target_h, 1))
        c_mal.font = Font(size=8, color="6b7280")
        c_mal.border = border
        c_mal.alignment = Alignment(horizontal="right", vertical="center")
        c_mal.number_format = "0.0"

        c_saldo = ws.cell(row, len(days) + 4, round(saldo, 1))
        saldo_color = "16a34a" if abs(saldo) <= 4 else "dc2626"
        c_saldo.font = Font(size=9, bold=True, color=saldo_color)
        c_saldo.border = border
        c_saldo.alignment = Alignment(horizontal="right", vertical="center")
        c_saldo.number_format = "+0.0;-0.0;0.0"

    # ── Bemanningsrad ────────────────────────────────────────────────────────
    bemanning_row = DATA_START + len(emp_rows)
    ws.row_dimensions[bemanning_row].height = 14

    c = ws.cell(bemanning_row, 1, "Bemanning")
    c.font = Font(bold=True, size=8, color="374151")
    c.fill = PatternFill("solid", fgColor=STAFFING_BG)
    c.border = border
    c.alignment = Alignment(horizontal="left", vertical="center")

    for i, d in enumerate(days):
        col = i + 2
        dag = staffing_dag[d]
        kval = staffing_kval[d]
        cell = ws.cell(bemanning_row, col)
        cell.fill = PatternFill("solid", fgColor=STAFFING_BG)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=7)
        if dag > 0 or kval > 0:
            parts = []
            if dag > 0:  parts.append(f"{dag}d")
            if kval > 0: parts.append(f"{kval}k")
            cell.value = " ".join(parts)

    for j in range(3):
        c = ws.cell(bemanning_row, len(days) + 2 + j)
        c.fill = PatternFill("solid", fgColor=STAFFING_BG)
        c.border = border

    # ── Frys namnkolumn och rubrikrader ──────────────────────────────────────
    ws.freeze_panes = f"B{DATA_START}"

    # ── Returnera som nedladdning ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schema_{group}_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
