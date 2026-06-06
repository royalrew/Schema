# -*- coding: utf-8 -*-
"""
FastAPI-router för Sintari RAG-Assistent.
Hanterar säkra schemasökningar och lokala manual-sökningar med strikta skyddsbarriärer (Guardrails).
"""

import os
import re
from datetime import date
from typing import Optional, Any
from pydantic import BaseModel, Field
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.db_models import EmployeeRow, SchedulePeriodRow, UserRow, RagDocumentRow
from app.auth_utils import get_current_user, get_current_user_optional

router = APIRouter(prefix="/api/chat", tags=["chat"])

# ── Pydantic-modeller ────────────────────────────────────────────────────────

class ChatMessageRequest(BaseModel):
    """Förfrågan som skickas från klienten vid ett chattmeddelande."""
    message: str = Field(..., min_length=1, description="Frågan eller meddelandet från användaren")


class ShiftDetailResponse(BaseModel):
    """Strukturerat skift-kort för Generative UI."""
    employee_name: str
    date_str: str
    shift_type: str
    start_time: str
    end_time: str
    is_unbooked: bool
    group_name: str
    label: str


class ChatResponse(BaseModel):
    """Respons som returneras till klienten."""
    response: str = Field(..., description="Svaret i klartext på svenska")
    category: str = Field("general", description="Kategori: general | error | schedule | manual | blocked")
    shift_details: Optional[ShiftDetailResponse] = Field(None, description="Strukturerade pass-detaljer för Generative UI")

# ── NLP & Hjälpfunktioner ───────────────────────────────────────────────────

# Röda zonen-regex för att fånga lagstiftning, OB-löner, fackliga tvister och GDPR-hälsa
BLOCK_KEYWORDS = [
    r"\bfack(en|et|lig|liga|smed|smeder)?\b",
    r"\bkollektivavtal\b",
    r"\barbetstidslag(en)?\b",
    r"\blagar\b",
    r"\blagligt\b",
    r"\brättigheter\b",
    r"\bob-ersättning(en)?\b",
    r"\bob-tillägg\b",
    r"\b(månads)?lön(en|er)?\b",
    r"\bövertid(ersättning)?\b",
    r"\bkränkt\b",
    r"\bsjukdom\b",
    r"\bsjukskriven\b",
    r"\bdiagnos\b",
    r"\bmedicinska\b",
]

MONTH_MAP = {
    "januari": 1, "jan": 1,
    "februari": 2, "feb": 2,
    "mars": 3, "mar": 3,
    "april": 4, "apr": 4,
    "maj": 5,
    "juni": 6, "jun": 6,
    "juli": 7, "jul": 7,
    "augusti": 8, "aug": 8,
    "september": 9, "sep": 9,
    "oktober": 10, "okt": 10,
    "november": 11, "nov": 11,
    "december": 12, "dec": 12
}


def _check_blocked_intent(message: str) -> Optional[str]:
    """
    Granskar om användarens fråga berör den "röda zonen" (lagar, fack, lön eller sjukdomsdetaljer).
    Returnerar ett artigt standardsvar om det är blockerat, annars None.
    """
    lower_message = message.lower()
    
    # GDPR Hälso-koll: fråga om någons sjukdom
    if ("sjuk" in lower_message or "sjukdom" in lower_message) and any(word in lower_message for word in ["varför", "anledning", "orsak"]):
        return (
            "Jag har tyvärr inte tillgång till medicinska uppgifter, sjukdomsorsaker eller "
            "känslig personlig information på grund av sekretessregler och GDPR."
        )

    # Allmän lag/fack/lön-koll
    for pattern in BLOCK_KEYWORDS:
        if re.search(pattern, lower_message):
            return (
                "Jag kan tyvärr inte svara på frågor om arbetsrätt, lagstiftning, löner, "
                "OB-ersättningar eller fackliga avtal. Vänligen vänd dig till din schemaansvarige, "
                "HR-avdelningen eller din fackliga representant."
            )
            
    return None


def _extract_text(file_path: str, ext: str) -> Optional[str]:
    """Extraherar text ur .txt/.md/.json/.pdf/.docx/.xlsx"""
    try:
        if ext in (".txt", ".md", ".json"):
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()

        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages) if pages else None

        if ext == ".docx":
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs) if paragraphs else None

        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows.append("  ".join(cells))
            return "\n".join(rows) if rows else None

    except Exception:
        pass
    return None


async def _search_rag_documents(query: str, db: AsyncSession) -> Optional[str]:
    """
    Söker i den lokala manualen SAMT i alla attesterade GRC-dokument i RAG-scopet.
    """
    query_words = [w.lower() for w in re.findall(r"\w+", query) if len(w) > 2]
    if not query_words:
        return None

    best_match_text = None
    best_score = 0.0
    source_title = ""

    # 1. Sök i systemmanualen först
    try:
        manual_path = os.path.join(os.path.dirname(__file__), "..", "resources", "user_manual.md")
        if os.path.exists(manual_path):
            with open(manual_path, "r", encoding="utf-8") as f:
                manual_content = f.read()
            sections = manual_content.split("\n## ")
            for sec in sections:
                score = 0
                sec_lower = sec.lower()
                for word in query_words:
                    if word in sec_lower:
                        score += 1
                if score > best_score and score >= 2:
                    best_score = float(score)
                    best_match_text = "## " + sec.strip()
                    source_title = "Systemmanualen"
    except Exception:
        pass

    # 2. Sök i attesterade dokument från databasen
    try:
        stmt = select(RagDocumentRow).where(RagDocumentRow.is_attested == True)
        result = await db.execute(stmt)
        docs = result.scalars().all()

        rag_dir = os.path.join(os.path.dirname(__file__), "..", "resources", "rag_docs")

        for doc in docs:
            file_path = os.path.join(rag_dir, doc.filename)
            if not os.path.exists(file_path):
                continue

            ext = os.path.splitext(doc.filename)[1].lower()
            content = _extract_text(file_path, ext)
            if not content:
                continue

            # Dela upp dokumentet i stycken för att ge specifika svar
            paragraphs = [p.strip() for p in content.split("\n\n") if p.strip()]
            for p in paragraphs:
                score = 0
                p_lower = p.lower()
                for word in query_words:
                    if word in p_lower:
                        score += 1
                # Ge lite högre prioritet åt specifika uppladdade dokument om de matchar bra
                adjusted_score = float(score) + 0.5
                if adjusted_score > best_score and score >= 2:
                    best_score = adjusted_score
                    best_match_text = p
                    source_title = doc.title
    except Exception:
        pass

    if best_match_text:
        return f"Här är informationen jag hittade i **{source_title}**:\n\n{best_match_text}"

    return None


def _extract_date_and_name(message: str) -> tuple[Optional[str], Optional[date]]:
    """
    Försöker extrahera ett medarbetarnamn och ett specifikt datum från användarens meddelande.
    Antar 2026 som standardår då det är schemaperiodens aktiva år.
    """
    lower_message = message.lower()
    
    # 1. Extrahera datum
    extracted_date = None
    
    # Format: YYYY-MM-DD
    iso_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", lower_message)
    if iso_match:
        try:
            extracted_date = date.fromisoformat(iso_match.group(0))
        except ValueError:
            pass
            
    if not extracted_date:
        # Format: 25:e maj, 25e maj, 25 maj, 25/5
        month_found = None
        day_found = None
        
        # Sök efter månad
        for m_name, m_num in MONTH_MAP.items():
            if m_name in lower_message:
                month_found = m_num
                break
                
        # Sök efter dagnummer i anslutning till månad eller fristående
        if month_found:
            # Letar efter t.ex. "25" eller "25:e" eller "25e" före eller efter månadsnamnet
            day_match = re.search(r"(\d{1,2})(?:e|:e|a|:a)?\s+" + list(MONTH_MAP.keys())[list(MONTH_MAP.values()).index(month_found)], lower_message)
            if not day_match:
                day_match = re.search(list(MONTH_MAP.keys())[list(MONTH_MAP.values()).index(month_found)] + r"\s+(\d{1,2})", lower_message)
            if day_match:
                day_found = int(day_match.group(1))
        else:
            # Sök efter t.ex. 25/5
            slash_match = re.search(r"(\d{1,2})/(\d{1,2})", lower_message)
            if slash_match:
                day_found = int(slash_match.group(1))
                month_found = int(slash_match.group(2))
                
        if day_found and month_found:
            try:
                # Vi standardiserar på år 2026 då schemat i Töreboda ligger där
                extracted_date = date(2026, month_found, day_found)
            except ValueError:
                pass

    # 2. Extrahera namn (mycket enkel matchning)
    # Vi söker efter ord med stor bokstav i originalmeddelandet, eller vanliga namn
    # Undvik vanliga ord i början av meningen
    extracted_name = None
    
    # Exkludera kända stoppord för namn
    stop_names = {
        "hur", "jobbar", "fredag", "måndag", "tisdag", "onsdag", "torsdag", "lördag", "söndag",
        "den", "maj", "juni", "juli", "vem", "visa", "hämta", "sök", "kolla", "se", "schema",
        "arbetspass", "period", "grupp", "för", "på", "till", "med", "av", "om", "eller", "men",
        "och", "har", "här", "var", "när", "jag", "min", "mitt", "du", "din", "ditt"
    }
    
    words = re.findall(r"\b[A-Za-zÅåÄäÖöé\-]+\b", message)
    for word in words:
        w_lower = word.lower()
        if w_lower in stop_names:
            continue
        # Om ordet börjar med stor bokstav eller om vi bara tar första icke-stoppordet som ser ut som ett namn
        if word[0].isupper() and len(word) >= 3:
            extracted_name = word
            break
            
    if not extracted_name:
        # Fallback: ta första bästa ord som är längre än 3 bokstäver och inte stoppord
        for word in words:
            w_lower = word.lower()
            if w_lower not in stop_names and len(word) >= 3:
                extracted_name = word
                break

    return extracted_name, extracted_date

# ── Slutpunkter ──────────────────────────────────────────────────────────────

@router.post("", response_model=ChatResponse)
async def chat_interaction(
    req: ChatMessageRequest,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[UserRow] = Depends(get_current_user_optional),
):
    """
    Hanterar en konversationsfråga från användaren.
    Använder Guardrails och slår upp i databas eller manual.
    """
    message = req.message.strip()
    
    # 1. Granska Röda Zonen (Guardrails)
    blocked_response = _check_blocked_intent(message)
    if blocked_response:
        return ChatResponse(
            response=blocked_response,
            category="blocked"
        )
        
    # 2. Analysera om det är en schemafråga
    name_query, date_query = _extract_date_and_name(message)
    
    # Om vi hittade ett datum och en person (eller om personalen frågar "när jobbar jag")
    is_self_query = "jag" in message.lower() or "min" in message.lower() or "mitt" in message.lower()
    
    if (name_query or is_self_query) and date_query:
        if not current_user:
            return ChatResponse(
                response="För att söka efter arbetspass eller visa scheman behöver du logga in på ditt konto. Klicka på Kundlogin för att fortsätta.",
                category="blocked"
            )
        # Hitta medarbetare
        emp_row = None
        if is_self_query:
            if not current_user.employee_id:
                return ChatResponse(
                    response="Ditt användarkonto är inte kopplat till någon anställd profil. Kontakta administratören.",
                    category="error"
                )
            stmt = select(EmployeeRow).where(EmployeeRow.id == current_user.employee_id)
            emp_row = (await db.execute(stmt)).scalar_one_or_none()
        else:
            # Sök efter medarbetare med liknande namn
            stmt = select(EmployeeRow).where(EmployeeRow.name.ilike(f"%{name_query}%"))
            emp_row = (await db.execute(stmt)).scalars().first()

        if not emp_row:
            return ChatResponse(
                response=f"Jag hittade tyvärr ingen medarbetare som matchar '{name_query}' i systemet.",
                category="error"
            )

        # RBAC Säkerhetskontroll: Personal får enbart fråga om sig själva!
        if current_user.role == "personal" and emp_row.id != current_user.employee_id:
            return ChatResponse(
                response="Du har tyvärr inte behörighet att se detaljerade arbetspass för andra medarbetare.",
                category="blocked"
            )

        # Hitta schema för perioden
        stmt = select(SchedulePeriodRow).where(
            SchedulePeriodRow.group_name == emp_row.group_name,
            SchedulePeriodRow.year == date_query.year,
            SchedulePeriodRow.month == date_query.month,
        )
        period_row = (await db.execute(stmt)).scalar_one_or_none()
        
        date_str = date_query.isoformat()
        
        if not period_row or not period_row.schedule:
            return ChatResponse(
                response=f"Det finns inget genererat eller publicerat schema för {emp_row.name}s grupp ({emp_row.group_name}) i {date_query.strftime('%B %Y')}.",
                category="schedule"
            )
            
        # Sök efter skiftet i schemalistan
        day_entry = None
        for day in period_row.schedule:
            if day.get("employee_id") == emp_row.id and day.get("date") == date_str:
                day_entry = day
                break
                
        if not day_entry:
            return ChatResponse(
                response=f"Jag hittade inget inplanerat pass för {emp_row.name} den {date_str}.",
                category="schedule"
            )
            
        shift = day_entry.get("shift")
        absence = day_entry.get("absence")
        
        if absence:
            abs_type = absence.get("absence_type", "Ledig")
            return ChatResponse(
                response=f"{emp_row.name} jobbar inte den {date_str} på grund av registrerad frånvaro: **{abs_type}**.",
                category="schedule"
            )
            
        if not shift:
            return ChatResponse(
                response=f"{emp_row.name} är **ledig** den {date_str}.",
                category="schedule"
            )
            
        # Bygg Generative UI respons
        shift_type = shift.get("shift_type", "DAG")
        is_unbooked = bool(shift.get("is_unbooked", False))
        
        # Hämta start/sluttider
        segments = shift.get("segments", [])
        start_time = "07:00"
        end_time = "16:00"
        if segments:
            # Extrahera tid "HH:MM" från ISO datetime
            s_dt = segments[0].get("start_time", "")
            e_dt = segments[0].get("end_time", "")
            if "T" in s_dt:
                start_time = s_dt.split("T")[1][:5]
            if "T" in e_dt:
                end_time = e_dt.split("T")[1][:5]
                
        label_text = "Dagpass"
        if shift_type == "dag_tidig":
            label_text = "Morgonpass (tidig)"
        elif shift_type == "kval_kort":
            label_text = "Kort kvällspass"
        elif shift_type == "kval_lang":
            label_text = "Långt kvällspass"
        elif shift_type == "delad_tur":
            label_text = "Delad tur"
        elif shift_type == "natt":
            label_text = "Nattpass"
        elif shift_type == "obokad":
            label_text = "Obokad flex-tid"

        text_response = f"{emp_row.name} jobbar **{label_text}** ({shift_type.upper()}) på {date_str} mellan kl. **{start_time}–{end_time}** i grupp **{emp_row.group_name}**."
        if is_unbooked:
            text_response += " Detta är inplanerat som obokad fyllnadstid för att nå kontraktstimmar."

        details = ShiftDetailResponse(
            employee_name=emp_row.name,
            date_str=date_str,
            shift_type=shift_type,
            start_time=start_time,
            end_time=end_time,
            is_unbooked=is_unbooked,
            group_name=emp_row.group_name,
            label=label_text
        )
        
        return ChatResponse(
            response=text_response,
            category="schedule",
            shift_details=details
        )

    # 3. Sök i RAG-dokument (Navigations-assistenten + uppladdade dokument)
    manual_response = await _search_rag_documents(message, db)
    if manual_response:
        return ChatResponse(
            response=manual_response,
            category="manual"
        )

    # 4. Standard fallback
    is_greeting = any(word in message.lower() for word in ["hej", "hallå", "tjena", "hejsan", "hello", "hi", "start"])
    
    if not current_user:
        if is_greeting or len(message) < 5:
            return ChatResponse(
                response=(
                    "Hej! Jag är din Sintari-assistent. Jag kan hjälpa dig med:\n"
                    "• **Allmänna frågor:** Fråga t.ex. *'Hur fungerar schemamotorn?'* eller *'Går det att integrera systemet?'*\n"
                    "• **Säkerhet & GDPR:** Sök t.ex. på *'Hur skyddas medarbetardata?'* eller *'Är systemet GDPR-säkrat?'*\n\n"
                    "Vad vill du veta mer om?"
                ),
                category="general"
            )
        else:
            return ChatResponse(
                response=(
                    "Jag förstår din fråga, men som Sintari-assistent är jag begränsad till att svara på frågor "
                    "om schemaläggning, GDPR-säkerhet, externa integrationer och vårt schemasystem. Jag kan tyvärr inte svara "
                    "på frågor utanför dessa ämnesområden.\n\n"
                    "Fråga mig gärna om hur schemamotorn fungerar, hur vi skyddar din data eller om våra integrationer!"
                ),
                category="general"
            )

    # Inloggade användare
    if is_greeting or len(message) < 5:
        org_name = os.getenv("ORGANIZATION_NAME", "Töreboda Hemvård")
        return ChatResponse(
            response=(
                f"Hej! Jag är din Sintari-assistent för {org_name}. Jag kan hjälpa dig med:\n"
                "• **Schemafrågor:** Fråga t.ex. *'När jobbar jag den 25 maj?'* eller *'Hur jobbar Elin på fredag?'*\n"
                "• **Navigeringshjälp:** Sök t.ex. på *'Hur lägger jag in önskeschema?'* eller *'Var ser jag timsaldo?'*\n\n"
                "Vad vill du ha hjälp med?"
            ),
            category="general"
        )
    else:
        return ChatResponse(
            response=(
                "Jag förstår din fråga, men min kunskap är begränsad till verksamhetens schema, "
                "passtider, timsaldon och hur du navigerar i systemet. Jag kan tyvärr inte svara på allmänna "
                "frågor utanför detta.\n\n"
                "Ställ gärna en fråga relaterad till ditt schema, semestrar eller passtider så hjälper jag dig!"
            ),
            category="general"
        )
